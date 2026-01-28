from datetime import datetime, timedelta
from dataclasses import dataclass
from io import BytesIO
import asyncio

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import pandas as pd
import matplotlib.pyplot as plt

from aiogram import Bot
from aiogram.types import BufferedInputFile

from src.bitrix.api.bitrix import BitrixAPI
from src.classes.cls_const import TaskRole, StageType
from src.db.database import BitrixDB
from src.db.models import TaskUser, Stage


@dataclass
class ExportInfo:
    creator: str
    executor: str
    task_bit_id: int
    task_title: str
    group: str
    stage: str
    created: datetime
    deadline: datetime | None = None
    test_date: datetime | None = None
    closed: datetime | None = None


class TaskExport:
    def __init__(self, db: BitrixDB):
        self.db = db

    async def schedule_send(self, chat_id: int | str, bot: Bot, run_hour: int, run_minute: int = 0):
        async def sleep_until(target_hour: int, target_minute: int = 0):
            now = datetime.now()
            target_time = now.replace(hour=target_hour, minute=target_minute, second=0, microsecond=0)
            if now >= target_time:
                target_time += timedelta(days=1)
            sleep_duration = (target_time - now).total_seconds()
            await asyncio.sleep(sleep_duration)

        while True:
            await sleep_until(run_hour, run_minute)
            await self.send_stat(chat_id, bot)

    async def send_stat(self, chat_id: int | str, bot: Bot):
        methods = (
            (self.create_and_closed_tasks, 30),
            (self.creator_stat, 30),
            (self.executor_stat, "m"),
            (self.save_png, 1),
            (self.queue_png, "d")
        )

        t = datetime.now()
        groups = await self.db.get_task_group(analytics=True)
        for group in groups:
            for method, day_delta in methods:
                try:
                    if isinstance(day_delta, int):
                        start = t - timedelta(days=day_delta)
                    elif day_delta == "m":
                        start = datetime(t.year, t.month, 1)

                    buf = await method(start=start, end=t, group_id=group.id)  # noqa
                    if isinstance(buf, BytesIO):
                        await bot.send_photo(
                            chat_id, BufferedInputFile(buf.read(), f"{method.__name__}.png")
                        )
                    elif isinstance(buf, BufferedInputFile):
                        await bot.send_document(chat_id, buf)

                except Exception as e:
                    print(f"error TaskExport -> {method.__name__}: {e}")

    async def get_tasks(self, start: datetime, end: datetime, group_id: int = None) -> dict[str, list[ExportInfo]]:
        result: dict[str, list[ExportInfo]] = {}
        tasks = await self.db.get_closed_tasks(start=start, end=end, group_id=group_id)

        all_stages = await self.db.get_task_stage(group_id=group_id)
        stages_ids = []
        if all_stages[0].group.max_tasks:
            for i in range(len(all_stages)):
                if all_stages[i].in_queue:
                    stages_ids = [s.id for s in all_stages[i:-1]]
                    break
        else:
            stages_ids = [s.id for s in all_stages]
        tasks += await self.db.get_tasks_with(stage_ids=stages_ids)

        for task in tasks:
            creator, executor = "None", "None"
            for task_user in task.task_users:  # type: TaskUser
                if task_user.role == TaskRole.CREATOR:
                    creator = task_user.user.full_name
                elif task_user.role == TaskRole.EXECUTOR:
                    executor = task_user.user.full_name

            if executor not in result:
                result[executor] = []

            result[executor].append(
                ExportInfo(
                    creator=creator,
                    executor=executor,
                    task_bit_id=task.bit_task_id,
                    task_title=task.title,
                    group=task.group.title,
                    stage=task.stage.title if task.stage else "None",
                    created=task.created_date,
                    deadline=task.deadline,
                    test_date=task.test_date,
                    closed=task.closed_date
                )
            )
        return result

    @staticmethod
    def calc_expired_time(task: ExportInfo) -> tuple:
        expired_test_time = ""
        expired_close_time = ""

        if task.deadline:
            if task.closed and task.closed > task.deadline:
                expired_close_time = (task.closed - task.deadline).total_seconds() // 3600
            elif (task.deadline < datetime.now()) and (not task.closed):
                expired_close_time = (datetime.now() - task.deadline).total_seconds() // 3600

            # test time
            if task.test_date and task.test_date > task.deadline:
                expired_test_time = (task.test_date - task.deadline).total_seconds() // 3600
            elif (task.deadline < datetime.now()) and (not task.test_date) and (not task.closed):
                expired_test_time = (datetime.now() - task.deadline).total_seconds() // 3600

        return expired_test_time, expired_close_time

    async def save_xlsx(self, start: datetime, end: datetime, group_id: int) -> BytesIO:
        def apply_border(cell_):
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell_.border = border

        def apply_conditional_formatting(row_number, deadline: datetime, column: int):
            if deadline and deadline < datetime.now():
                fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                ws.cell(row=row_number, column=column).fill = fill

        # create title -----------------------------------------------------------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Задачи"
        columns = [
            "Исполнитель", "Статус", "Задача", "Заказчик",
            "Срок", "На тестирование", "Создан", "Принят"
        ]
        columns_width = [30, 20, 40, 30, 16, 16, 16, 16]
        ws.append(columns)
        for i, col in enumerate(columns):
            cell = ws.cell(row=1, column=i + 1)
            cell.font = Font(bold=True)
            col_letter = get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = columns_width[i]
            apply_border(cell)
        # ------------------------------------------------------------------------------------------

        tasks: list[ExportInfo] = []
        t = await self.get_tasks(start, end, group_id=group_id)
        for e_tasks in t.values():
            tasks += e_tasks

        for task in tasks:
            expired_test_time, expired_close_time = self.calc_expired_time(task)

            row = [
                task.executor, task.stage, f"{task.task_bit_id} - {task.task_title}", task.creator,
                task.deadline.strftime("%d.%m.%Y %H:%M") if task.deadline else "-",
                task.test_date.strftime("%d.%m.%Y %H:%M") if task.test_date else "-",
                task.created.strftime("%d.%m.%Y %H:%M") if task.created else "-",
                task.closed.strftime("%d.%m.%Y %H:%M") if task.closed else "-",

            ]
            ws.append(row)
            if expired_test_time:
                apply_conditional_formatting(len(ws["A"]), task.deadline, 8)

            if expired_close_time:
                apply_conditional_formatting(len(ws["A"]), task.deadline, 9)

            # Apply border to all cells in the row
            for i, value in enumerate(row):
                cell = ws.cell(row=len(ws["A"]), column=i + 1)
                apply_border(cell)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    async def save_png(self, start: datetime, end: datetime, group_id: int) -> BytesIO | None:
        tasks: list[ExportInfo] = []
        t = await self.get_tasks(start, end, group_id=group_id)
        if not t:
            return None
        for e_tasks in t.values():
            tasks += e_tasks

        data = []
        for task in tasks:
            # expired_test_time, expired_close_time = self.calc_expired_time(task)

            row = [
                task.executor if len(task.executor) < 31 else task.executor[:29] + ".",
                task.stage if len(task.stage) < 21 else task.stage[:19] + ".",
                f"{task.task_bit_id} - {task.task_title if len(task.task_title) < 30 else task.task_title[:29] + '.'}",
                task.creator if len(task.creator) < 31 else task.creator[:29] + ".",
                task.deadline.strftime("%d.%m.%Y %H:%M") if task.deadline else "-",
                task.test_date.strftime("%d.%m.%Y %H:%M") if task.test_date else "-",
                task.created.strftime("%d.%m.%Y %H:%M") if task.closed else "-",
                task.closed.strftime("%d.%m.%Y %H:%M") if task.closed else "-"
            ]
            data.append(row)

        df = pd.DataFrame(
            data,
            columns=[
                "Исполнитель", "Статус", "Задача", "Заказчик",
                "Срок", "На тестирование", "Создан", "Принят"
            ]
        )

        fig, ax = plt.subplots(figsize=(19, len(df) * 0.4 + 2))  # уменьшаем размер фигуры
        ax.axis("tight")
        ax.axis("off")

        # Добавляем заголовок
        title = f"Отчет по задачам {tasks[0].group} {end.strftime('%d.%m.%Y')}"
        plt.figtext(0.5, 0.95, title, ha="center", va="top", fontsize=14, fontweight="bold")

        table = ax.table(
            cellText=df.values, colLabels=df.columns, cellLoc="center", loc="center", bbox=[0.05, 0.2, 0.9, 0.6]
        )
        table.auto_set_font_size(False)
        table.set_fontsize(10)

        # Устанавливаем ширину столбцов
        col_widths = [0.15, 0.1, 0.3, 0.15, 0.1, 0.1, 0.1, 0.1]  # пример ширины столбцов в пропорциях
        for i, width in enumerate(col_widths):
            table.auto_set_column_width(i)  # отключаем автоматическое выставление ширины столбцов
            for j in range(len(df) + 1):
                cell = table[(j, i)]
                cell.set_width(width)

        # Устанавливаем жирный шрифт для заголовков столбцов
        for (i, j), cell in table.get_celld().items():
            if i == 0:
                cell.set_text_props(weight="bold")

            if j == len(df.columns) - 2 and i > 0 and df.iloc[i - 1, -2]:  # columns[-2]
                cell.set_facecolor("#CCCCFF")

            if j == len(df.columns) - 1 and i > 0 and df.iloc[i - 1, -1]:  # columns[-1]
                cell.set_facecolor("#FFCCCC")

        plt.tight_layout()

        buffer = BytesIO()
        plt.savefig(buffer, format="png", dpi=300)  # увеличиваем DPI для лучшего качества
        buffer.seek(0)
        return buffer

    async def create_and_closed_tasks(self, start: datetime, end: datetime, group_id: int) -> BytesIO | None:
        stages = await self.db.get_task_stage(group_id=group_id)
        all_created = await self.db.get_stage_task_counts(stage_ids=[s.id for s in stages[1:]])
        all_closed = await self.db.get_stage_task_counts(stage_ids=[stages[-1].id])
        all_closed += await self.db.get_stage_task_counts(
            stage_ids=[s.id for s in stages if s.stage_type == StageType.TESTING]
        )

        # in period
        created_tasks = await self.db.get_created_tasks(start, end, group_id=group_id)
        created_tasks = [i for i in created_tasks if i.stage_id != stages[0].id]
        closed_tasks = await self.db.get_closed_tasks(start, end, group_id=group_id)

        last_create = all_created - len(created_tasks)
        last_closed = all_closed - len(closed_tasks)

        last_create_num = last_create  # to fill in missing dates
        last_closed_num = last_closed  # to fill in missing dates

        created_days = {}
        closed_days = {}

        for created_task in created_tasks:
            task_created_date = created_task.created_date.strftime("%d.%m.%Y")
            last_create += 1
            created_days[task_created_date] = last_create

        for closed_task in closed_tasks:
            task_closed_date = closed_task.closed_date.strftime("%d.%m.%Y")
            last_closed += 1
            closed_days[task_closed_date] = last_closed

        if (not closed_tasks) and (not created_tasks):
            return None

        # to fill in missing dates --------------------------------------------------
        all_dates = pd.date_range(start=start, end=end).strftime("%d.%m.%Y").tolist()
        all_created_date = {}
        all_closed_date = {}
        for date in all_dates:
            if date in created_days:
                last_create_num = created_days[date]

            if date in closed_days:
                last_closed_num = closed_days[date]

            if last_create_num or last_closed_num:
                all_created_date[date] = last_create_num
                all_closed_date[date] = last_closed_num

        # Plot the data
        plt.figure(figsize=(10, 5))
        plt.plot(
            list(all_created_date.keys()), list(all_created_date.values()),
            label="Создано", color="blue", marker="o"
        )
        plt.plot(
            list(all_closed_date.keys()), list(all_closed_date.values()),
            label="Завершено", color="green", marker="o"
        )

        # Formatting the plot
        group_task = created_tasks[0] if created_tasks else closed_tasks[0]
        plt.xlabel("Дата")
        plt.ylabel("Количество задач")
        plt.title(f"Динамика созданных и закрытых задач {group_task.group.title}")
        plt.xticks(rotation=45)
        plt.legend()
        plt.grid(True)
        plt.tight_layout()

        # Save the plot to a BytesIO buffer
        buffer = BytesIO()
        plt.savefig(buffer, format="png", dpi=300)
        buffer.seek(0)

        return buffer

    async def get_creator_stat(self, start: datetime, end: datetime, group_id: int) -> list | None:
        tasks = await self.db.get_created_tasks(start, end, group_id=group_id)
        creators = {}
        for task in tasks:
            for t_user in task.task_users:  # type: TaskUser
                if t_user.role == TaskRole.CREATOR:
                    if t_user.user.full_name not in creators:
                        creators[t_user.user.full_name] = 1
                    else:
                        creators[t_user.user.full_name] += 1
                    break

        if not creators:
            return None

        return sorted(creators.items(), key=lambda item: item[1], reverse=True)

    async def creator_stat(self, start: datetime, end: datetime, group_id: int) -> BytesIO | None:
        sorted_creators = await self.get_creator_stat(start, end, group_id)
        if not sorted_creators:
            return None
        group = await self.db.get_task_group(id_=group_id)

        # Prepare data for the pie chart
        top_creators = sorted_creators[:9]
        other_creators = sorted_creators[9:]

        top_labels = [f"{creator[0]} - {creator[1]}" for creator in top_creators]
        top_sizes = [creator[1] for creator in top_creators]

        if other_creators:
            other_label = "Другие"
            other_size = sum([creator[1] for creator in other_creators])
            top_labels.append(f"{other_label} - {other_size}")
            top_sizes.append(other_size)

        # Plot the pie chart
        fig, ax = plt.subplots(figsize=(8, 8))
        ax.pie(top_sizes, labels=top_labels, autopct="%1.1f%%", startangle=140)
        ax.set_title(
            f"Задачи {group[0].title}, созданные авторами "
            f"c {start.strftime('%d.%m.%Y')} по {end.strftime('%d.%m.%Y')}"
        )

        # Save the plot to a BytesIO buffer
        buffer = BytesIO()
        plt.savefig(buffer, format="png", bbox_inches="tight", dpi=300)
        buffer.seek(0)

        return buffer

    async def executor_stat(self, start: datetime, end: datetime, group_id: int) -> BytesIO | None:
        all_tasks = await self.db.get_closed_tasks(start, end, group_id=group_id)
        executors: dict = {}
        executors_tasks: dict[str, set] = {}
        for task in all_tasks:
            if task.paid:
                continue

            for t_user in task.task_users:  # type: TaskUser
                if t_user.role in (TaskRole.EXECUTOR, TaskRole.CO_EXECUTOR):
                    if t_user.user.full_name not in executors:
                        executors[t_user.user.full_name] = [0, 0]
                        executors_tasks[t_user.user.full_name] = set()

                    if task.id not in executors_tasks[t_user.user.full_name]:
                        executors[t_user.user.full_name][0] += (task.allocated_time or 0)
                        executors[t_user.user.full_name][1] += 1
                        executors_tasks[t_user.user.full_name].add(task.id)

        if not executors:
            return None

        # Создание DataFrame из данных
        data = {
            'Executor': [
                (" ".join(i.split()[:2]) if i.index(" ") else i) + f" - {executors[i][1]}" for i in executors.keys()
            ],
            'Hours': [val[0] // 3600 for val in executors.values()],
            'Tasks': [val[1] for val in executors.values()]
        }
        df = pd.DataFrame(data)

        # Преобразование меток оси x в числовые значения
        x = range(len(df))
        hours = df['Hours']
        tasks = df['Tasks']
        labels = df['Executor']
        width = 0.35  # Ширина гистограммы

        fig, ax = plt.subplots(figsize=(10, 6))

        # Гистограмма для количества часов
        ax.bar([i - width / 2 for i in x], hours, width, label='Количество часов')

        # Гистограмма для количества задач
        ax.bar([i + width / 2 for i in x], tasks, width, label='Количество задач')

        # Добавление подписей и заголовков
        ax.set_ylabel("Количество")
        ax.set_title(f"{all_tasks[0].group.title} количество часов и задач по исполнителям "
                     f"c {start.strftime('%d.%m.%Y')} по {end.strftime('%d.%m.%Y')}")
        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=60, ha="right")
        ax.legend()

        plt.tight_layout()

        # Сохранение графика в буфер
        buffer = BytesIO()
        plt.savefig(buffer, format="png")
        buffer.seek(0)

        return buffer

    async def get_task_stage_time(self, bitrix: BitrixAPI, group_id: int, closed_days: int = 0, queue: bool = True):
        def apply_border(cell_):
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell_.border = border

        # create title -----------------------------------------------------------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Задачи"
        columns = [
            "id", "Задача", "Заказчик", "Менеджер", "Исполнитель",
            "Создано", "Изменено", "Статус",
            "Последний комментарий", "Дата п. комментария", "Автор п. комментария"
        ]
        columns_width = [
            8, 40, 30, 30, 30,
            20, 20, 20,
            40, 20, 30
        ]

        all_stages: dict[str, float] = {}
        all_stages_obj: dict[str, Stage] = {}
        additional_stages = (StageType.ERROR, StageType.TESTING, StageType.WAIT, StageType.DEVELOP)
        queue_stages_id = []

        stages = await self.db.get_task_stage(group_id=group_id)
        for stage in stages:
            if (queue is False) or stage.in_queue or stage.stage_type in additional_stages:
                if stage.id != stages[-1].id:
                    queue_stages_id.append(stage.id)

            all_stages[stage.title] = 0.0
            all_stages_obj[stage.title] = stage
            columns.append(stage.title)
            columns_width.append(10)

        ws.append(columns)
        for i, col in enumerate(columns):
            cell = ws.cell(row=1, column=i + 1)
            cell.font = Font(bold=True)
            col_letter = get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = columns_width[i]
            apply_border(cell)
        # ------------------------------------------------------------------------------------------

        tasks = []
        now = datetime.now()
        if queue_stages_id:
            tasks += await self.db.get_tasks_with(stage_ids=queue_stages_id)

        if closed_days:
            tasks += await self.db.get_closed_tasks(start=now-timedelta(days=closed_days), end=now, group_id=group_id)

        if not tasks:
            return None

        for task in tasks:
            task_users = self.db.sort_task_roles(task.task_users)
            commet = await self.db.get_comment(task.id)
            row = [
                task.bit_task_id, task.title,
                task_users.creator.user.__str__() if task_users.creator else "",
                task_users.manager.user.__str__() if task_users.manager else "",
                task_users.executor.user.__str__() if task_users.executor else "",

                task.created_date.strftime("%Y.%m.%d %H:%M"),
                "last_change_date",
                task.stage.title,

                commet[-1].text if commet else "-",
                commet[-1].created_date.strftime("%Y.%m.%d %H:%M") if commet else "-",
                commet[-1].user.full_name if commet else "-",
            ]

            # calc -------------------------------
            history = await bitrix.get_history(task.bit_task_id, stage=True)
            history = history.get("list")

            last_stage = ""
            info = all_stages.copy()
            previous_date: datetime = task.created_date
            for stage_history in history:
                date = datetime.fromisoformat(stage_history["createdDate"]).astimezone().replace(tzinfo=None)
                f: str = stage_history["value"]["from"]
                last_stage = stage_history["value"]["to"]

                was_time = (date - previous_date).total_seconds()
                if f and f in info.keys() and was_time > 60:
                    if all_stages_obj[f].stage_type in (StageType.DEVELOP, StageType.WAIT):
                        info[f] = round(info[f] + was_time / 3600, 1)
                    else:
                        info[f] = round(was_time / 3600, 1)

                previous_date = date

            else:
                if last_stage and last_stage in info.keys():
                    was_time = (now - previous_date).total_seconds()
                    if was_time > 60:
                        if all_stages_obj[last_stage].stage_type in (StageType.DEVELOP, StageType.WAIT):
                            info[last_stage] = round(info[last_stage] + was_time / 3600, 1)
                        else:
                            info[last_stage] = round(was_time / 3600, 1)

            # last change
            if history:
                stage_date = datetime.fromisoformat(history[-1]["createdDate"]).astimezone().replace(tzinfo=None)
            else:
                stage_date = task.created_date

            commet_date = commet[-1].created_date if commet else task.created_date
            last_change_time = stage_date if stage_date > commet_date else commet_date
            row[6] = last_change_time.strftime("%Y.%m.%d %H:%M")

            row += info.values()

            ws.append(row)
            # Apply border to all cells in the row
            for i, value in enumerate(row):
                cell = ws.cell(row=len(ws["A"]), column=i + 1)
                apply_border(cell)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    async def get_fifo_queue(self, start: datetime, end: datetime, group_id: int) -> BufferedInputFile:
        def apply_border(cell_):
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell_.border = border

        # create title -----------------------------------------------------------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Задачи"
        columns = ["id", "Задача", "Заказчик", "Менеджер", "Дата в очереди"]
        columns_width = [10, 40, 30, 30, 20]
        ws.append(columns)
        for i, col in enumerate(columns):
            cell = ws.cell(row=1, column=i + 1)
            cell.font = Font(bold=True)
            col_letter = get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = columns_width[i]
            apply_border(cell)
        # ------------------------------------------------------------------------------------------

        tasks_fifo = await self.db.get_fifo_queue(group_id=group_id, limit=None)
        for task in tasks_fifo:
            task_users = self.db.sort_task_roles(task.task_users)
            row = [
                task.bit_task_id, task.title, task_users.creator.user.full_name,
                task_users.manager.user.full_name if task_users.manager else "",
                task.queue_date.strftime("%Y.%m.%d %H:%M:%S")
            ]
            ws.append(row)
            # Apply border to all cells in the row
            for i, value in enumerate(row):
                cell = ws.cell(row=len(ws["A"]), column=i + 1)
                apply_border(cell)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return BufferedInputFile(buffer.read(), f"{tasks_fifo[0].group.title}_fifo.xlsx")

    async def queue_png(self, start: datetime, end: datetime, group_id: int) -> BytesIO | None:
        data: list[list] = []

        tasks_fifo = await self.db.get_fifo_queue(group_id=group_id, limit=20)
        for id_, task in enumerate(tasks_fifo, start=1):
            task_users = self.db.sort_task_roles(task.task_users)
            row = [
                id_, task.bit_task_id,
                task.title if len(task.title) < 60 else task.title[:59] + '.',
                task_users.creator.user.full_name,
                task_users.manager.user.full_name if task_users.manager else "",
                task.queue_date.strftime("%Y.%m.%d %H:%M:%S")
            ]
            data.append(row)

        if data:
            return self.png_table(
                f"Очередь задач для {tasks_fifo[0].group.title}",
                ["№", "id", "Задача", "Заказчик", "Менеджер", "Дата в очереди"],
                [0.1, 0.1, 0.1, 0.1, 0.1, 0.1],
                data
            )

    @staticmethod
    async def s_queue_png(db, group_id: int) -> BytesIO | None:
        data: list[list] = []

        tasks_fifo = await db.get_fifo_queue(group_id=group_id, limit=20)
        for id_, task in enumerate(tasks_fifo, start=1):
            task_users = db.sort_task_roles(task.task_users)
            row = [
                id_, task.bit_task_id,
                task.title if len(task.title) < 60 else task.title[:59] + '.',
                task_users.creator.user.full_name,
                task_users.manager.user.full_name if task_users.manager else "",
                task.queue_date.strftime("%Y.%m.%d %H:%M:%S")
            ]
            data.append(row)

        if data:
            return TaskExport.png_table(
                f"Очередь задач для {tasks_fifo[0].group.title}",
                ["№", "id", "Задача", "Заказчик", "Менеджер", "Дата в очереди"],
                [0.1, 0.1, 0.1, 0.1, 0.1, 0.1],
                data
            )

    @staticmethod
    def png_table(title: str, columns: list[str], col_widths: list[float], data: list[list]) -> BytesIO | None:
        df = pd.DataFrame(data, columns=columns)

        fig, ax = plt.subplots(figsize=(19, len(df) * 0.4 + 2))  # уменьшаем размер фигуры
        ax.axis("tight")
        ax.axis("off")

        # Добавляем заголовок
        plt.figtext(0.5, 0.95, title, ha="center", va="top", fontsize=14, fontweight="bold")

        table = ax.table(
            cellText=df.values, colLabels=df.columns, cellLoc="center", loc="center", bbox=[0.05, 0.05, 0.9, 0.85]
        )
        table.auto_set_font_size(False)
        table.set_fontsize(10)

        # Устанавливаем ширину столбцов
        for i, width in enumerate(col_widths):
            table.auto_set_column_width(i)  # отключаем автоматическое выставление ширины столбцов
            for j in range(len(df) + 1):
                cell = table[(j, i)]
                cell.set_width(width)

        # Устанавливаем жирный шрифт для заголовков столбцов
        for (i, j), cell in table.get_celld().items():
            if i == 0:
                cell.set_text_props(weight="bold")

        plt.tight_layout()
        buffer = BytesIO()
        plt.savefig(buffer, format="png", dpi=300)  # увеличиваем DPI для лучшего качества
        buffer.seek(0)
        return buffer
